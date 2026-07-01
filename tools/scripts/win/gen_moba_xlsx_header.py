"""
生成 MOBA pixel_moba 数值表的 xlsx 表头(Row1-5)+ 示例数据。

使用:
    python tools/scripts/win/gen_moba_xlsx_header.py

产物:
    bizconfig/static-xlsx/moba/pixel_moba/{unit,hero_profiles,hero_level_bonus,
    skill,skill_level,effect,item}.xlsx

注意:
    - .proto schema 由 biz_export 工具从 xlsx 表头自动生成,无需手写
    - 此脚本仅生成表头与 1-2 行示例数据,数值留空由策划填充
"""

from openpyxl import Workbook
from pathlib import Path
import sys

OUTPUT_DIR = Path(r"d:\git-project\BeastServer-project\bizconfig\static-xlsx\moba\pixel_moba")


# ---------- 公共属性字段(在 unit 和 hero_level_bonus 中复用) ----------
# 顺序与字段名都对齐,用于 Row1
COMBAT_ATTRS = [
    # (字段名, 类型, 中文注释, 默认值, 可见性)
    ("max_hp",                "int32",  "1级最大生命",          600,   "!s!c"),
    ("max_mana",              "int32",  "1级最大法力",          300,   "!s!c"),
    ("hp_regen",              "float",  "每秒回血",            2.5,   "!s!c"),
    ("mana_regen",            "float",  "每秒回蓝",            1.5,   "!s!c"),
    ("physical_attack",       "int32",  "1级物理攻击",         55,    "!s!c"),
    ("magic_attack",          "int32",  "1级法术攻击",         0,     "!s!c"),
    ("physical_defense",      "int32",  "1级物理防御",         20,    "!s!c"),
    ("magic_defense",         "int32",  "1级法术防御",         15,    "!s!c"),
    ("attack_before",         "float",  "攻击前摇(秒)",        0.3,   "!s!c"),
    ("attack_after",          "float",  "攻击后摇(秒)",        0.4,   "!s!c"),
    ("attack_interval",       "float",  "攻击间隔(秒)",        1.4,   "!s!c"),
    ("attack_range",          "float",  "攻击距离(像素)",      40,    "!s!c"),
    ("move_speed",            "float",  "移动速度(像素/秒)",   220,   "!s!c"),
    ("collision_radius",      "float",  "碰撞半径(像素)",      16,    "!s!c"),
    ("vision_range",          "float",  "视野范围(像素)",       400,   "!s!c"),
    ("physical_lifesteal",    "float",  "物理吸血(0-1)",        0.0,   "!s!c"),
    ("magic_lifesteal",       "float",  "法术吸血(0-1)",        0.0,   "!s!c"),
    ("crit_rate",             "float",  "暴击率(0-1)",          0.0,   "!s!c"),
    ("crit_damage",           "float",  "暴击伤害倍率(1.5=150%)", 2.0, "!s!c"),
    ("physical_pen",          "int32",  "物理穿透(固定值)",    0,     "!s!c"),
    ("physical_pen_pct",      "float",  "物理穿透(0-1)",        0.0,   "!s!c"),
    ("magic_pen",             "int32",  "法术穿透(固定值)",    0,     "!s!c"),
    ("magic_pen_pct",         "float",  "法术穿透(0-1)",        0.0,   "!s!c"),
    ("cd_reduction",          "float",  "冷却缩减(0-0.4)",       0.0,   "!s!c"),
    ("dodge",                 "float",  "闪避率(0-1)",          0.0,   "!s!c"),
    ("hit_rate",              "float",  "命中率(0-1)",          1.0,   "!s!c"),
    ("slow_resist",           "float",  "减速抗性(0-1)",        0.0,   "!s!c"),
    ("tenacity",              "float",  "韧性(减控0-1)",        0.0,   "!s!c"),
]


def write_sheet(ws, columns, sample_rows):
    """
    columns: List[Tuple[field_name, type_str, constraint, visibility, comment]]
        Row1 = name
        Row2 = type_str
        Row3 = constraint
        Row4 = visibility
        Row5 = comment
    sample_rows: List[Dict[field_name -> value]]
    """
    for col_idx, (name, type_str, constraint, vis, comment) in enumerate(columns, start=1):
        ws.cell(row=1, column=col_idx, value=name)
        ws.cell(row=2, column=col_idx, value=type_str)
        ws.cell(row=3, column=col_idx, value=constraint)
        ws.cell(row=4, column=col_idx, value=vis)
        ws.cell(row=5, column=col_idx, value=comment)

    for row_offset, sample in enumerate(sample_rows, start=6):
        for col_idx, (name, type_str, *_rest) in enumerate(columns, start=1):
            val = sample.get(name, "")
            # Excel 把 Python bool 转成大写 TRUE/FALSE,biz_export 校验只接受小写 true/false
            if isinstance(val, bool):
                val = "true" if val else "false"
            # openpyxl 把整数值的 float 降级为 int,写回 xlsx 后 excelize 读取字符串变化
            # (如 0.0 -> "0" 而非 "0.0"),强制 float 类型保持一致
            elif type_str in ("float", "double") and isinstance(val, (int, float)):
                val = float(val)
            ws.cell(row=row_offset, column=col_idx, value=val)


# ============================================================
# 1. unit 表 — 单位基础属性(英雄/小兵/防御塔/野怪共表)
# ============================================================
def gen_unit_xlsx():
    wb = Workbook()
    ws = wb.active
    ws.title = "Unit"

    columns = [
        ("id",                  "uint32", "notnull&only", "!s!c", "标识ID(1xxx英雄/2xxx小兵/3xxx塔/4xxx野怪)"),
        ("index",               "string", "notnull&only", "!s!c", "唯一字符串索引,如 warrior/archer/melee_minion"),
        ("name",                "string", "notnull",      "!s!c", "显示名"),
        ("unit_type",           "string", "notnull",      "!s!c", "单位类型: hero/minion/tower/creep"),
        ("is_neutral",          "bool",   "optional",     "!s",   "是否中立(野怪用),默认 false"),
        ("spawn_interval_ms",   "int32",  "optional",     "!s",   "刷新间隔毫秒(小兵用),其他留空"),
        ("death_reward_gold",   "int32",  "notnull",      "!s!c", "击杀奖励金币"),
        ("death_reward_exp",    "int32",  "notnull",      "!s!c", "击杀奖励经验"),
        ("client_scene_path",    "string", "optional",     "!c",   "客户端场景资源路径"),
        ("client_icon_path",    "string", "optional",     "!c",   "客户端图标路径"),
    ]

    # 接上公共战斗属性
    for name, type_str, comment, _default, vis in COMBAT_ATTRS:
        constraint = "notnull"
        columns.append((name, type_str, constraint, vis, comment))

    # 示例:战士(1001) + 法师(1002) + 近战兵(2001) + 一塔(3001)
    samples = [
        {"id": 1001, "index": "warrior", "name": "战士", "unit_type": "hero",
         "is_neutral": False, "spawn_interval_ms": "", "death_reward_gold": 300, "death_reward_exp": 100,
         "client_scene_path": "res://gameplay/content/heroes/1001_warrior/hero.tscn",
         "client_icon_path": "res://gameplay/content/heroes/1001_warrior/icon.png",
         "max_hp": 600, "max_mana": 300, "hp_regen": 2.5, "mana_regen": 1.5,
         "physical_attack": 55, "magic_attack": 0, "physical_defense": 20, "magic_defense": 15,
         "attack_before": 0.3, "attack_after": 0.4, "attack_interval": 1.4,
         "attack_range": 40, "move_speed": 220,
         "collision_radius": 16, "vision_range": 400,
         "physical_lifesteal": 0.0, "magic_lifesteal": 0.0, "crit_rate": 0.0, "crit_damage": 2.0,
         "physical_pen": 0, "physical_pen_pct": 0.0, "magic_pen": 0, "magic_pen_pct": 0.0,
         "cd_reduction": 0.0, "dodge": 0.0, "hit_rate": 1.0, "slow_resist": 0.0, "tenacity": 0.0},

        {"id": 1002, "index": "mage", "name": "法师", "unit_type": "hero",
         "is_neutral": False, "spawn_interval_ms": "", "death_reward_gold": 300, "death_reward_exp": 100,
         "client_scene_path": "res://gameplay/content/heroes/1002_mage/hero.tscn",
         "client_icon_path": "res://gameplay/content/heroes/1002_mage/icon.png",
         "max_hp": 450, "max_mana": 400, "hp_regen": 1.5, "mana_regen": 2.0,
         "physical_attack": 30, "magic_attack": 70, "physical_defense": 14, "magic_defense": 18,
         "attack_before": 0.25, "attack_after": 0.35, "attack_interval": 1.6,
         "attack_range": 200, "move_speed": 210,
         "collision_radius": 16, "vision_range": 420,
         "physical_lifesteal": 0.0, "magic_lifesteal": 0.0, "crit_rate": 0.0, "crit_damage": 2.0,
         "physical_pen": 0, "physical_pen_pct": 0.0, "magic_pen": 0, "magic_pen_pct": 0.0,
         "cd_reduction": 0.0, "dodge": 0.0, "hit_rate": 1.0, "slow_resist": 0.0, "tenacity": 0.0},

        {"id": 1003, "index": "archer", "name": "弓箭手", "unit_type": "hero",
         "is_neutral": False, "spawn_interval_ms": "", "death_reward_gold": 300, "death_reward_exp": 100,
         "client_scene_path": "res://gameplay/content/heroes/1003_archer/hero.tscn",
         "client_icon_path": "res://gameplay/content/heroes/1003_archer/icon.png",
         "max_hp": 500, "max_mana": 250, "hp_regen": 2.0, "mana_regen": 1.5,
         "physical_attack": 60, "magic_attack": 0, "physical_defense": 16, "magic_defense": 14,
         "attack_before": 0.2, "attack_after": 0.3, "attack_interval": 1.2,
         "attack_range": 180, "move_speed": 230,
         "collision_radius": 16, "vision_range": 440,
         "physical_lifesteal": 0.0, "magic_lifesteal": 0.0, "crit_rate": 0.0, "crit_damage": 2.0,
         "physical_pen": 0, "physical_pen_pct": 0.0, "magic_pen": 0, "magic_pen_pct": 0.0,
         "cd_reduction": 0.0, "dodge": 0.0, "hit_rate": 1.0, "slow_resist": 0.0, "tenacity": 0.0},

        {"id": 2001, "index": "melee_minion", "name": "近战兵", "unit_type": "minion",
         "is_neutral": False, "spawn_interval_ms": 30000, "death_reward_gold": 20, "death_reward_exp": 10,
         "client_scene_path": "res://gameplay/content/minions/melee_minion.tscn",
         "client_icon_path": "",
         "max_hp": 200, "max_mana": 0, "hp_regen": 0, "mana_regen": 0,
         "physical_attack": 12, "magic_attack": 0, "physical_defense": 5, "magic_defense": 5,
         "attack_before": 0.35, "attack_after": 0.35, "attack_interval": 1.7,
         "attack_range": 30, "move_speed": 150,
         "collision_radius": 12, "vision_range": 200,
         "physical_lifesteal": 0.0, "magic_lifesteal": 0.0, "crit_rate": 0.0, "crit_damage": 2.0,
         "physical_pen": 0, "physical_pen_pct": 0.0, "magic_pen": 0, "magic_pen_pct": 0.0,
         "cd_reduction": 0.0, "dodge": 0.0, "hit_rate": 1.0, "slow_resist": 0.0, "tenacity": 0.0},

        {"id": 3001, "index": "tower_t1", "name": "一塔", "unit_type": "tower",
         "is_neutral": False, "spawn_interval_ms": "", "death_reward_gold": 250, "death_reward_exp": 0,
         "client_scene_path": "res://gameplay/content/towers/tower_t1.tscn",
         "client_icon_path": "",
         "max_hp": 2000, "max_mana": 0, "hp_regen": 0, "mana_regen": 0,
         "physical_attack": 100, "magic_attack": 0, "physical_defense": 50, "magic_defense": 50,
         "attack_before": 0.1, "attack_after": 0.1, "attack_interval": 1.0,
         "attack_range": 250, "move_speed": 0,
         "collision_radius": 24, "vision_range": 600,
         "physical_lifesteal": 0.0, "magic_lifesteal": 0.0, "crit_rate": 0.0, "crit_damage": 2.0,
         "physical_pen": 0, "physical_pen_pct": 0.0, "magic_pen": 0, "magic_pen_pct": 0.0,
         "cd_reduction": 0.0, "dodge": 0.0, "hit_rate": 1.0, "slow_resist": 0.0, "tenacity": 0.0},
    ]

    write_sheet(ws, columns, samples)
    wb.save(OUTPUT_DIR / "unit.xlsx")
    print(f"OK  unit.xlsx  ({len(columns)} cols, {len(samples)} samples)")


# ============================================================
# 2. hero_profiles 表 — 英雄特化
# ============================================================
def gen_hero_profiles_xlsx():
    wb = Workbook()
    ws = wb.active
    ws.title = "HeroProfiles"

    columns = [
        ("hero_id",                "uint32", "notnull&only", "!s!c", "外键 -> unit.id"),
        ("class",                  "string", "notnull",      "!s!c", "英雄定位: archer/warrior/mage"),
        ("max_level",              "int32",  "notnull",      "!s!c", "最大等级(固定 10)"),
        ("start_gold",             "int32",  "notnull",      "!s!c", "初始金币"),
        ("respawn_base_sec",       "int32",  "notnull",      "!s!c", "基础复活时间秒"),
        ("respawn_per_level_sec",  "float",  "notnull",      "!s!c", "每死亡等级加复活秒数"),
        ("skill_q_id",             "uint32", "notnull",      "!s!c", "Q技能ID,外键 -> skill.id"),
        ("skill_w_id",             "uint32", "notnull",      "!s!c", "W技能ID,外键 -> skill.id"),
        ("skill_e_id",             "uint32", "notnull",      "!s!c", "E技能ID,外键 -> skill.id"),
        ("skill_r_id",             "uint32", "notnull",      "!s!c", "R大招ID,外键 -> skill.id"),
        ("r_unlock_level",         "int32",  "notnull",      "!s!c", "R大招解锁等级(通常 6)"),
        ("passive_skill_id",       "uint32", "optional",     "!s!c", "被动技能ID,外键 -> skill.id,无则留空"),
        ("base_attack_bonus",      "float",  "optional",     "!s",   "基础攻击加成(远程/近战修正),仅服务端"),
        ("lore",                   "string", "optional",     "!c",   "英雄背景故事(客户端 tooltip)"),
    ]

    samples = [
        {"hero_id": 1001, "class": "warrior", "max_level": 10, "start_gold": 500,
         "respawn_base_sec": 5, "respawn_per_level_sec": 2.0,
         "skill_q_id": 5001, "skill_w_id": 5002, "skill_e_id": 5003, "skill_r_id": 5004,
         "r_unlock_level": 6, "passive_skill_id": 5005, "base_attack_bonus": 0.0,
         "lore": "近战战士,擅长冲锋与控制"},
        {"hero_id": 1002, "class": "mage", "max_level": 10, "start_gold": 500,
         "respawn_base_sec": 5, "respawn_per_level_sec": 2.0,
         "skill_q_id": 5011, "skill_w_id": 5012, "skill_e_id": 5013, "skill_r_id": 5014,
         "r_unlock_level": 6, "passive_skill_id": 5015, "base_attack_bonus": 0.0,
         "lore": "远程法师,高爆发魔法伤害"},
        {"hero_id": 1003, "class": "archer", "max_level": 10, "start_gold": 500,
         "respawn_base_sec": 5, "respawn_per_level_sec": 2.0,
         "skill_q_id": 5021, "skill_w_id": 5022, "skill_e_id": 5023, "skill_r_id": 5024,
         "r_unlock_level": 6, "passive_skill_id": 5025, "base_attack_bonus": 0.0,
         "lore": "远程弓手,持续物理输出"},
    ]

    write_sheet(ws, columns, samples)
    wb.save(OUTPUT_DIR / "hero_profiles.xlsx")
    print(f"OK  hero_profiles.xlsx  ({len(columns)} cols, {len(samples)} samples)")


# ============================================================
# 3. hero_level_bonus 表 — 英雄每级加成(每英雄 10 行,存总属性)
# ============================================================
def gen_hero_level_bonus_xlsx():
    wb = Workbook()
    ws = wb.active
    ws.title = "HeroLevelBonus"

    columns = [
        ("hero_id",         "uint32", "notnull",     "!s!c", "外键 -> hero_profiles.hero_id"),
        ("level",           "int32",  "notnull",     "!s!c", "等级 1-10"),
        ("skill_point",     "int32",  "notnull",     "!s!c", "升级时获得的技能点(通常 1,R解锁级可给 2)"),
        ("exp_to_next",     "int32",  "notnull",     "!s!c", "升到下一级所需经验(level 10 留 0)"),
        ("unlock_slot",     "string", "optional",     "!s!c", "本等级解锁内容(如 R 表示解锁大招),无则留空"),
    ]

    # 复用 COMBAT_ATTRS(本等级总值)
    for name, type_str, comment, _default, vis in COMBAT_ATTRS:
        # 注释加上"(该等级总值)"前缀,提示策划填总值而非加成
        columns.append((name, type_str, "notnull", vis, comment + "(该等级总值)"))

    # 示例:warrior 的 1-3 级(其他英雄/其他等级由策划补全)
    samples = [
        {"hero_id": 1001, "level": 1,  "skill_point": 1, "exp_to_next": 100, "unlock_slot": "",
         "max_hp": 600, "max_mana": 300, "hp_regen": 2.5, "mana_regen": 1.5,
         "physical_attack": 55, "magic_attack": 0, "physical_defense": 20, "magic_defense": 15,
         "attack_before": 0.3, "attack_after": 0.4, "attack_interval": 1.4,
         "attack_range": 40, "move_speed": 220,
         "collision_radius": 16, "vision_range": 400,
         "physical_lifesteal": 0.0, "magic_lifesteal": 0.0, "crit_rate": 0.0, "crit_damage": 2.0,
         "physical_pen": 0, "physical_pen_pct": 0.0, "magic_pen": 0, "magic_pen_pct": 0.0,
         "cd_reduction": 0.0, "dodge": 0.0, "hit_rate": 1.0, "slow_resist": 0.0, "tenacity": 0.0},

        {"hero_id": 1001, "level": 2,  "skill_point": 1, "exp_to_next": 150, "unlock_slot": "",
         "max_hp": 680, "max_mana": 320, "hp_regen": 2.7, "mana_regen": 1.6,
         "physical_attack": 62, "magic_attack": 0, "physical_defense": 22, "magic_defense": 16,
         "attack_before": 0.3, "attack_after": 0.4, "attack_interval": 1.39,
         "attack_range": 40, "move_speed": 220,
         "collision_radius": 16, "vision_range": 400,
         "physical_lifesteal": 0.0, "magic_lifesteal": 0.0, "crit_rate": 0.0, "crit_damage": 2.0,
         "physical_pen": 0, "physical_pen_pct": 0.0, "magic_pen": 0, "magic_pen_pct": 0.0,
         "cd_reduction": 0.0, "dodge": 0.0, "hit_rate": 1.0, "slow_resist": 0.0, "tenacity": 0.0},

        {"hero_id": 1001, "level": 6,  "skill_point": 2, "exp_to_next": 600, "unlock_slot": "R",
         "max_hp": 1050, "max_mana": 420, "hp_regen": 3.5, "mana_regen": 2.0,
         "physical_attack": 90, "magic_attack": 0, "physical_defense": 30, "magic_defense": 22,
         "attack_before": 0.28, "attack_after": 0.38, "attack_interval": 1.25,
         "attack_range": 40, "move_speed": 225,
         "collision_radius": 16, "vision_range": 420,
         "physical_lifesteal": 0.0, "magic_lifesteal": 0.0, "crit_rate": 0.0, "crit_damage": 2.0,
         "physical_pen": 0, "physical_pen_pct": 0.0, "magic_pen": 0, "magic_pen_pct": 0.0,
         "cd_reduction": 0.0, "dodge": 0.0, "hit_rate": 1.0, "slow_resist": 0.0, "tenacity": 0.0},
    ]

    write_sheet(ws, columns, samples)
    wb.save(OUTPUT_DIR / "hero_level_bonus.xlsx")
    print(f"OK  hero_level_bonus.xlsx  ({len(columns)} cols, {len(samples)} samples)")


# ============================================================
# 4. skill 表 — 技能 metadata
# ============================================================
def gen_skill_xlsx():
    wb = Workbook()
    ws = wb.active
    ws.title = "Skill"

    columns = [
        ("id",                  "uint32", "notnull&only", "!s!c", "技能ID(5xxx段)"),
        ("index",               "string", "notnull&only", "!s!c", "唯一字符串索引,如 archer_q/warrior_r"),
        ("name",                "string", "notnull",      "!s!c", "显示名"),
        ("skill_type",          "string", "notnull",      "!s!c", "类型: active(主动)/passive(被动)/aura(光环)"),
        ("cast_type",           "string", "notnull",      "!s!c", "施法方式: instant/projectile/point/target_aoe"),
        ("damage_type",         "string", "notnull",      "!s!c", "伤害类型: physical/magical/true(真实伤害)"),
        ("max_level",           "int32",  "notnull",      "!s!c", "最大等级(固定 9)"),
        ("default_effect_ids",  "string", "optional",     "!s",   "触发effect id链,JSON数组如 [1,2],仅服务端"),
        ("client_icon_path",    "string", "optional",     "!c",   "客户端图标路径"),
        ("desc",                "string", "optional",     "!c",   "客户端描述(支持{level}占位符)"),
        ("mana_type",           "string", "notnull",      "!s!c", "消耗类型: mana/health/none(无消耗)"),
    ]

    samples = [
        {"id": 5001, "index": "warrior_q", "name": "冲锋斩", "skill_type": "active",
         "cast_type": "instant", "damage_type": "physical", "max_level": 9,
         "default_effect_ids": "[1]", "client_icon_path": "res://gameplay/content/skills/warrior_q/icon.png",
         "desc": "向目标方向冲锋,造成{level}点物理伤害", "mana_type": "mana"},
        {"id": 5011, "index": "mage_q", "name": "火球术", "skill_type": "active",
         "cast_type": "projectile", "damage_type": "magical", "max_level": 9,
         "default_effect_ids": "[2]", "client_icon_path": "res://gameplay/content/skills/mage_q/icon.png",
         "desc": "发射火球,造成{level}点法术伤害", "mana_type": "mana"},
        {"id": 5021, "index": "archer_q", "name": "穿透箭", "skill_type": "active",
         "cast_type": "projectile", "damage_type": "physical", "max_level": 9,
         "default_effect_ids": "[3]", "client_icon_path": "res://gameplay/content/skills/archer_q/icon.png",
         "desc": "射出穿透箭,对路径上敌人造成{level}点物理伤害", "mana_type": "mana"},
    ]

    write_sheet(ws, columns, samples)
    wb.save(OUTPUT_DIR / "skill.xlsx")
    print(f"OK  skill.xlsx  ({len(columns)} cols, {len(samples)} samples)")


# ============================================================
# 5. skill_level 表 — 技能每级数值(每技能 9 行)
# ============================================================
def gen_skill_level_xlsx():
    wb = Workbook()
    ws = wb.active
    ws.title = "SkillLevel"

    columns = [
        ("skill_id",           "uint32", "notnull",      "!s!c", "外键 -> skill.id"),
        ("level",              "int32",  "notnull",      "!s!c", "等级 1-9"),
        ("cooldown_ms",        "int32",  "notnull",      "!s!c", "冷却时间毫秒"),
        ("mana_cost",          "int32",  "notnull",      "!s!c", "法力消耗(无消耗留 0)"),
        ("cast_range",         "float",  "notnull",      "!s!c", "施法距离(像素)"),
        ("radius",             "float",  "optional",     "!s",   "AOE 半径,无 AOE 留空,仅服务端"),
        ("damage_base",        "int32",  "optional",     "!s",   "基础伤害,代入 effect 公式,仅服务端"),
        ("heal_base",          "int32",  "optional",     "!s",   "基础治疗,仅服务端"),
        ("buff_duration_ms",   "int32",  "optional",     "!s",   "buff 持续时间,仅服务端"),
        ("projectile_speed",   "float",  "optional",     "!s",   "弹道速度(像素/秒),instant 留空,仅服务端"),
        ("physical_scaling",   "float",  "optional",     "!s",   "物攻加成比(0.5=50% 物攻),仅服务端"),
        ("magic_scaling",      "float",  "optional",     "!s",   "法攻加成比,仅服务端"),
        ("max_hp_scaling",     "float",  "optional",     "!s",   "最大生命加成比(部分技能用),仅服务端"),
        ("lifesteal_pct",      "float",  "optional",     "!s",   "该技能吸血率(0-1),仅服务端"),
    ]

    # 示例:warrior_q 的 1/2/9 级
    samples = [
        {"skill_id": 5001, "level": 1, "cooldown_ms": 8000, "mana_cost": 40, "cast_range": 100,
         "radius": "", "damage_base": 80, "heal_base": "", "buff_duration_ms": "",
         "projectile_speed": "", "physical_scaling": 0.5, "magic_scaling": 0.0,
         "max_hp_scaling": 0.0, "lifesteal_pct": 0.0},
        {"skill_id": 5001, "level": 2, "cooldown_ms": 7800, "mana_cost": 45, "cast_range": 100,
         "radius": "", "damage_base": 130, "heal_base": "", "buff_duration_ms": "",
         "projectile_speed": "", "physical_scaling": 0.55, "magic_scaling": 0.0,
         "max_hp_scaling": 0.0, "lifesteal_pct": 0.0},
        {"skill_id": 5001, "level": 9, "cooldown_ms": 6000, "mana_cost": 80, "cast_range": 120,
         "radius": "", "damage_base": 400, "heal_base": "", "buff_duration_ms": "",
         "projectile_speed": "", "physical_scaling": 0.8, "magic_scaling": 0.0,
         "max_hp_scaling": 0.0, "lifesteal_pct": 0.0},
    ]

    write_sheet(ws, columns, samples)
    wb.save(OUTPUT_DIR / "skill_level.xlsx")
    print(f"OK  skill_level.xlsx  ({len(columns)} cols, {len(samples)} samples)")


# ============================================================
# 6. effect 表 — 效果模板(只服务端可见)
# ============================================================
def gen_effect_xlsx():
    wb = Workbook()
    ws = wb.active
    ws.title = "Effect"

    columns = [
        ("id",            "uint32", "notnull&only", "!s!c", "效果ID(6xxx段)"),
        ("name",          "string", "notnull",      "!s!c", "效果显示名(客户端 tooltip 引用)"),
        ("effect_type",   "string", "notnull",      "!s",   "类型: damage/heal/buff_stun/buff_slow/teleport/spawn_unit,仅服务端"),
        ("formula",       "string", "notnull",      "!s",   "公式模板,$前缀变量: $attacker.physical_attack*0.5 + $skill_level.damage_base,仅服务端"),
        ("target_type",   "string", "notnull",      "!s",   "目标: self/enemy/ally/point/aoe,仅服务端"),
        ("element",        "string", "optional",     "!s",   "元素: physical/magical/true,留空默认 physical,仅服务端"),
        ("stack_max",     "int32",  "optional",     "!s",   "可叠加次数(buff 用),默认 1,仅服务端"),
        ("dispellable",   "bool",   "optional",     "!s",   "是否可被驱散,默认 true,仅服务端"),
    ]

    samples = [
        {"id": 6001, "name": "物理伤害", "effect_type": "damage", "formula": "$skill_level.damage_base + $attacker.physical_attack * $skill_level.physical_scaling",
         "target_type": "enemy", "element": "physical", "stack_max": "", "dispellable": ""},
        {"id": 6002, "name": "治疗", "effect_type": "heal", "formula": "$skill_level.heal_base + $attacker.magic_attack * $skill_level.magic_scaling",
         "target_type": "ally", "element": "true", "stack_max": "", "dispellable": ""},
        {"id": 6003, "name": "眩晕", "effect_type": "buff_stun", "formula": "$skill_level.buff_duration_ms",
         "target_type": "enemy", "element": "", "stack_max": 1, "dispellable": False},
        {"id": 6004, "name": "减速", "effect_type": "buff_slow", "formula": "0.3;$skill_level.buff_duration_ms",
         "target_type": "enemy", "element": "", "stack_max": 1, "dispellable": True},
    ]

    write_sheet(ws, columns, samples)
    wb.save(OUTPUT_DIR / "effect.xlsx")
    print(f"OK  effect.xlsx  ({len(columns)} cols, {len(samples)} samples)")


# ============================================================
# 7. item 表 — 商品(消耗品 + 装备无合成)
# ============================================================
def gen_item_xlsx():
    wb = Workbook()
    ws = wb.active
    ws.title = "Item"

    columns = [
        ("id",                "uint32", "notnull&only", "!s!c", "商品ID(7xxx段)"),
        ("index",             "string", "notnull&only", "!s!c", "唯一索引,如 hp_potion/boots/longsword"),
        ("name",              "string", "notnull",      "!s!c", "显示名"),
        ("item_type",         "string", "notnull",      "!s!c", "类型: consumable(消耗品)/equipment(装备)"),
        ("equip_slot",        "string", "optional",     "!s!c", "装备槽: weapon/helmet/boots/ring(消耗品留空)"),
        ("price",             "int32",  "notnull",      "!s!c", "价格"),
        ("is_consumable",     "bool",   "notnull",      "!s!c", "是否消耗(用一次消失)"),
        ("stack_max",         "int32",  "notnull",      "!s!c", "可堆叠上限(消耗品通常 3-9,装备 1)"),
        ("stat_bonus",        "string", "optional",     "!s!c", "属性加成,k=v;k=v 格式,如 max_hp=100;physical_attack=15"),
        ("effect_id",         "uint32", "optional",     "!s",   "使用效果ID,外键 -> effect.id(消耗品用),仅服务端"),
        ("sell_price_pct",    "float",  "notnull",      "!s!c", "出售价格比例(0-1,通常 0.6)"),
        ("client_icon_path",  "string", "optional",     "!c",   "客户端图标路径"),
    ]

    samples = [
        {"id": 7001, "index": "hp_potion", "name": "生命药水", "item_type": "consumable",
         "equip_slot": "", "price": 50, "is_consumable": True, "stack_max": 5,
         "stat_bonus": "", "effect_id": 6002, "sell_price_pct": 0.6,
         "client_icon_path": "res://gameplay/content/items/hp_potion/icon.png"},
        {"id": 7002, "index": "mana_potion", "name": "法力药水", "item_type": "consumable",
         "equip_slot": "", "price": 50, "is_consumable": True, "stack_max": 5,
         "stat_bonus": "", "effect_id": 6002, "sell_price_pct": 0.6,
         "client_icon_path": "res://gameplay/content/items/mana_potion/icon.png"},
        {"id": 7011, "index": "longsword", "name": "长剑", "item_type": "equipment",
         "equip_slot": "weapon", "price": 350, "is_consumable": False, "stack_max": 1,
         "stat_bonus": "physical_attack=15", "effect_id": "", "sell_price_pct": 0.7,
         "client_icon_path": "res://gameplay/content/items/longsword/icon.png"},
        {"id": 7012, "index": "boots", "name": "速度之靴", "item_type": "equipment",
         "equip_slot": "boots", "price": 300, "is_consumable": False, "stack_max": 1,
         "stat_bonus": "move_speed=45", "effect_id": "", "sell_price_pct": 0.7,
         "client_icon_path": "res://gameplay/content/items/boots/icon.png"},
    ]

    write_sheet(ws, columns, samples)
    wb.save(OUTPUT_DIR / "item.xlsx")
    print(f"OK  item.xlsx  ({len(columns)} cols, {len(samples)} samples)")


# ============================================================
# 8. map_arena 表 — 地图配置(8 sheet,数据来自 dev-arena1.json)
# ============================================================
def gen_map_arena_xlsx():
    """生成地图配置表(8 sheet,数据来自 dev-arena1.json)。"""
    import json
    wb = Workbook()
    arena_id = 1

    def _j(arr):
        return json.dumps(arr, separators=(",", ":"))

    # ---- Sheet 1: MapArena (地图元信息) ----
    ws = wb.active
    ws.title = "MapArena"
    write_sheet(ws, [
        ("id",      "uint32", "notnull&only", "!s!c", "地图ID"),
        ("index",   "string", "notnull&only", "!s!c", "唯一索引,如 dev_arena1"),
        ("name",    "string", "notnull",      "!s!c", "显示名"),
        ("width",   "int32",  "notnull",      "!s!c", "宽(格)"),
        ("height",  "int32",  "notnull",      "!s!c", "高(格)"),
        ("version", "int32",  "notnull",      "!s!c", "版本号"),
    ], [{"id": arena_id, "index": "dev_arena1", "name": "开发竞技场1",
         "width": 150, "height": 150, "version": 1}])

    # ---- Sheet 2: MapBase (基地:blue/red) ----
    ws = wb.create_sheet("MapBase")
    write_sheet(ws, [
        ("arena_id", "uint32", "notnull", "!s!c", "外键 -> MapArena.id"),
        ("team",     "string", "notnull", "!s!c", "blue/red"),
        ("rect",     "string", "notnull", "!s!c", "[x,y,w,h] JSON"),
        ("core",     "string", "notnull", "!s!c", "[x,y,w,h] JSON 核心建筑"),
        ("shop",     "string", "notnull", "!s!c", "[x,y,w,h] JSON 商店"),
        ("spawn",    "string", "notnull", "!s!c", "[x,y] JSON 出生点"),
    ], [
        {"arena_id": arena_id, "team": "blue",
         "rect": _j([60, 0, 30, 18]), "core": _j([72, 6, 6, 6]),
         "shop": _j([64, 10, 4, 4]), "spawn": _j([74, 14])},
        {"arena_id": arena_id, "team": "red",
         "rect": _j([60, 132, 30, 18]), "core": _j([72, 138, 6, 6]),
         "shop": _j([82, 136, 4, 4]), "spawn": _j([74, 134])},
    ])

    # ---- Sheet 3: MapLane (兵线) ----
    ws = wb.create_sheet("MapLane")
    write_sheet(ws, [
        ("arena_id", "uint32", "notnull", "!s!c", "外键 -> MapArena.id"),
        ("lane_id",  "string", "notnull", "!s!c", "left/right/mid"),
        ("rect",     "string", "notnull", "!s!c", "[x,y,w,h] JSON"),
        ("path",     "string", "notnull", "!s!c", "[[x,y],...] JSON 路径点"),
    ], [
        {"arena_id": arena_id, "lane_id": "left",  "rect": _j([40, 18, 6, 114]),
         "path": _j([[43, 18], [43, 64], [43, 86], [43, 132]])},
        {"arena_id": arena_id, "lane_id": "right", "rect": _j([104, 18, 6, 114]),
         "path": _j([[107, 18], [107, 64], [107, 86], [107, 132]])},
    ])

    # ---- Sheet 4: MapRiver (河道) ----
    ws = wb.create_sheet("MapRiver")
    write_sheet(ws, [
        ("arena_id", "uint32", "notnull&only", "!s!c", "外键 -> MapArena.id"),
        ("rect",     "string", "notnull",      "!s!c", "[x,y,w,h] JSON"),
    ], [{"arena_id": arena_id, "rect": _j([0, 64, 150, 22])}])

    # ---- Sheet 5: MapTower (防御塔) ----
    ws = wb.create_sheet("MapTower")
    write_sheet(ws, [
        ("arena_id", "uint32", "notnull",      "!s!c", "外键 -> MapArena.id"),
        ("tower_id", "string", "notnull&only", "!s!c", "如 blue_left"),
        ("team",     "string", "notnull",      "!s!c", "blue/red"),
        ("rect",     "string", "notnull",      "!s!c", "[x,y,w,h] JSON"),
    ], [
        {"arena_id": arena_id, "tower_id": "blue_left",  "team": "blue", "rect": _j([41, 40, 4, 4])},
        {"arena_id": arena_id, "tower_id": "blue_right", "team": "blue", "rect": _j([105, 40, 4, 4])},
        {"arena_id": arena_id, "tower_id": "red_left",   "team": "red",  "rect": _j([41, 106, 4, 4])},
        {"arena_id": arena_id, "tower_id": "red_right",  "team": "red",  "rect": _j([105, 106, 4, 4])},
    ])

    # ---- Sheet 6: MapWall (墙) ----
    ws = wb.create_sheet("MapWall")
    wall_rects = [
        [8, 24, 6, 18], [30, 40, 12, 4], [136, 24, 6, 18], [108, 40, 12, 4],
        [8, 108, 6, 18], [30, 106, 12, 4], [136, 108, 6, 18], [108, 106, 12, 4],
        [64, 68, 6, 14], [80, 68, 6, 14], [34, 58, 4, 18], [112, 58, 4, 18],
    ]
    write_sheet(ws, [
        ("arena_id",   "uint32", "notnull", "!s!c", "外键 -> MapArena.id"),
        ("wall_index", "string", "notnull", "!s!c", "如 w0/w1"),
        ("rect",       "string", "notnull", "!s!c", "[x,y,w,h] JSON"),
    ], [{"arena_id": arena_id, "wall_index": f"w{i}", "rect": _j(r)}
        for i, r in enumerate(wall_rects)])

    # ---- Sheet 7: MapBush (草丛) ----
    ws = wb.create_sheet("MapBush")
    bush_rects = [
        [52, 66, 10, 8], [88, 66, 10, 8], [36, 48, 8, 8],
        [104, 48, 8, 8], [36, 96, 8, 8], [104, 96, 8, 8],
    ]
    write_sheet(ws, [
        ("arena_id",    "uint32", "notnull", "!s!c", "外键 -> MapArena.id"),
        ("bush_index", "string", "notnull", "!s!c", "如 b0/b1"),
        ("rect",        "string", "notnull", "!s!c", "[x,y,w,h] JSON"),
    ], [{"arena_id": arena_id, "bush_index": f"b{i}", "rect": _j(r)}
        for i, r in enumerate(bush_rects)])

    # ---- Sheet 8: MapMonsterSpawn (野怪刷新点) ----
    ws = wb.create_sheet("MapMonsterSpawn")
    write_sheet(ws, [
        ("arena_id", "uint32", "notnull",      "!s!c", "外键 -> MapArena.id"),
        ("spawn_id", "string", "notnull&only", "!s!c", "如 blue_buff"),
        ("type",     "string", "notnull",      "!s!c", "外键 -> unit.index"),
        ("rect",     "string", "notnull",      "!s!c", "[x,y,w,h] JSON"),
    ], [
        {"arena_id": arena_id, "spawn_id": "blue_buff",  "type": "buff",  "rect": _j([22, 30, 3, 3])},
        {"arena_id": arena_id, "spawn_id": "blue_wolf",  "type": "wolf",  "rect": _j([16, 54, 2, 2])},
        {"arena_id": arena_id, "spawn_id": "blue_bird",  "type": "bird",  "rect": _j([34, 58, 2, 2])},
        {"arena_id": arena_id, "spawn_id": "blue_golem", "type": "golem", "rect": _j([24, 18, 2, 2])},
        {"arena_id": arena_id, "spawn_id": "red_buff",   "type": "buff",  "rect": _j([125, 117, 3, 3])},
        {"arena_id": arena_id, "spawn_id": "red_wolf",   "type": "wolf",  "rect": _j([132, 94, 2, 2])},
        {"arena_id": arena_id, "spawn_id": "red_bird",   "type": "bird",  "rect": _j([114, 90, 2, 2])},
        {"arena_id": arena_id, "spawn_id": "red_golem",  "type": "golem", "rect": _j([124, 130, 2, 2])},
        {"arena_id": arena_id, "spawn_id": "river_boss", "type": "boss",  "rect": _j([73, 72, 4, 4])},
    ])

    wb.save(OUTPUT_DIR / "map_arena.xlsx")
    print(f"OK  map_arena.xlsx  (8 sheets)")


# ============================================================
# 入口
# ============================================================
if __name__ == "__main__":
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    failed = []
    for gen_fn, name in [
        (gen_unit_xlsx, "unit"),
        (gen_hero_profiles_xlsx, "hero_profiles"),
        (gen_hero_level_bonus_xlsx, "hero_level_bonus"),
        (gen_skill_xlsx, "skill"),
        (gen_skill_level_xlsx, "skill_level"),
        (gen_effect_xlsx, "effect"),
        (gen_item_xlsx, "item"),
        (gen_map_arena_xlsx, "map_arena"),
    ]:
        try:
            gen_fn()
        except PermissionError as e:
            print(f"FAIL  {name}.xlsx  (PermissionError: 文件被占用,请关闭 Excel 后重试)")
            failed.append(name)
    print(f"\nDone. 8 xlsx attempted at {OUTPUT_DIR}")
    if failed:
        print(f"Failed: {failed}")
        sys.exit(1)
